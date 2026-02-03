# -*- coding: utf-8 -*-

"""
Excel Exporter
Exports test session data to formatted Excel files.

Updated to include extended session metadata:
- Ürün Bilgileri (Product Info)
- Yazılım Bilgileri (Software Info)
- Cihaz Kalibrasyonları (Device Calibrations)
- Oturum Bilgileri (Session Info)

NOTE: İstasyon and SİP are NOT included in Excel (UI-only fields)
"""
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from models.test_session import TestSession
from models.enums import TestStatus
import config
from utils.logger import setup_logger

logger = setup_logger(__name__)


class ExcelExporter:
    """
    Exports test session data to formatted Excel files.
    
    Features:
    - Extended session metadata (product, software, calibrations)
    - Step-by-step results with status, duration, and comments
    - Color-coded pass/fail status
    - Professional formatting
    - Auto-column sizing
    
    NOTE: İstasyon and SİP are UI-only fields and NOT included in Excel export.
    """
    
    # Excel styling constants
    HEADER_FILL = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
    
    SECTION_FILL = PatternFill(start_color="3949ab", end_color="3949ab", fill_type="solid")
    SECTION_FONT = Font(color="FFFFFF", bold=True, size=11)
    
    PASS_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    FAIL_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    NA_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    WARNING_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    EXPIRED_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    
    BORDER_THIN = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self):
        """Initialize the Excel exporter"""
        logger.info("ExcelExporter initialized (with extended metadata support)")
    
    def export_session(self, session: TestSession, output_path: str) -> bool:
        """
        Export test session to Excel file.
        
        Args:
            session: TestSession to export
            output_path: Full path where Excel file should be saved
            
        Returns:
            True if export successful, False otherwise
        """
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Sonuçları"
            
            # Write content
            current_row = 1
            current_row = self._write_header_section(ws, session, current_row)
            current_row += 2  # Add spacing
            current_row = self._write_results_table(ws, session, current_row)
            
            # Auto-adjust column widths
            self._auto_size_columns(ws)
            
            # Save file
            wb.save(output_path)
            
            logger.info(f"Excel file exported successfully to: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to export Excel file: {e}")
            return False
    
    def _write_header_section(self, ws, session: TestSession, start_row: int) -> int:
        """
        Write session metadata header section with extended information.
        
        Includes:
        - Report Title
        - Ürün Bilgileri (Product Info)
        - Yazılım Bilgileri (Software Info)
        - Cihaz Kalibrasyonları (Device Calibrations)
        - Oturum Bilgileri (Session Timing)
        
        NOTE: İstasyon and SİP are NOT included (UI-only fields)
        
        Returns:
            Next available row number
        """
        row = start_row
        
        # =====================================================================
        # Report Title
        # =====================================================================
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws[f'A{row}']
        cell.value = "TEST PROSEDÜRÜ RAPORU"
        cell.font = Font(size=16, bold=True, color="1a237e")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 2
        
        # Check if extended metadata is available
        metadata = session.metadata if hasattr(session, 'metadata') and session.metadata else None
        
        # =====================================================================
        # ÜRÜN BİLGİLERİ (Product Information)
        # =====================================================================
        row = self._write_section_header(ws, row, "ÜRÜN BİLGİLERİ")
        
        if metadata:
            product_data = [
                ("Stok No:", metadata.stok_no or session.stock_number),
                ("Opsiyonel Stok No:", metadata.opsiyonel_stok_no),
                ("Tanım:", metadata.tanim),
                ("TEU UDK:", metadata.teu_udk),
                ("Seri No:", metadata.seri_no or session.serial_number),
                ("261 Revizyonu:", metadata.revizyon_261),
                ("Test Donanımı Revizyon:", metadata.test_donanimi_revizyon),
                ("Test Yazılımı Revizyon:", metadata.test_yazilimi_revizyon),
                ("İş Tipi No:", metadata.is_tipi_no),
            ]
        else:
            # Fallback to basic session info
            product_data = [
                ("Stok No:", session.stock_number),
                ("Seri No:", session.serial_number),
            ]
        
        for label, value in product_data:
            if value:  # Only write non-empty values
                row = self._write_info_row(ws, row, label, value)
        
        row += 1  # Spacing
        
        # =====================================================================
        # YAZILIM BİLGİLERİ (Software Information)
        # =====================================================================
        if metadata and (metadata.kay_yazilimi_versiyon or metadata.sky_yazilimi_versiyon):
            row = self._write_section_header(ws, row, "YAZILIM BİLGİLERİ")
            
            software_data = [
                ("KAY Yazılımı Versiyon:", metadata.kay_yazilimi_versiyon),
                ("SKY Yazılımı Versiyon:", metadata.sky_yazilimi_versiyon),
            ]
            
            for label, value in software_data:
                if value:
                    row = self._write_info_row(ws, row, label, value)
            
            row += 1  # Spacing
        
        # =====================================================================
        # CİHAZ KALİBRASYONLARI (Device Calibrations)
        # =====================================================================
        if metadata:
            calibration_data = [
                ("FLUKE ESA620:", metadata.fluke_esa620_kalibrasyon, 'fluke_esa620_kalibrasyon'),
                ("ITALSEA 7PROGLCD:", metadata.italsea_7proglcd_kalibrasyon, 'italsea_7proglcd_kalibrasyon'),
                ("Geratech:", metadata.geratech_kalibrasyon, 'geratech_kalibrasyon'),
                ("IBA MagicMax:", metadata.iba_magicmax_kalibrasyon, 'iba_magicmax_kalibrasyon'),
                ("IBA Primus A:", metadata.iba_primus_a_kalibrasyon, 'iba_primus_a_kalibrasyon'),
            ]
            
            # Check if any calibration data exists
            has_calibrations = any(date_val for _, date_val, _ in calibration_data)
            
            if has_calibrations:
                row = self._write_section_header(ws, row, "CİHAZ KALİBRASYONLARI")
                
                for label, date_val, field_name in calibration_data:
                    if date_val:
                        # Format date for display
                        formatted_date = metadata.format_date_for_display(field_name)
                        status = metadata.get_calibration_status(field_name)
                        row = self._write_calibration_row(ws, row, label, formatted_date, status)
                
                row += 1  # Spacing
        
        # =====================================================================
        # OTURUM BİLGİLERİ (Session Timing) - Generic timing, NO İstasyon/SİP
        # =====================================================================
        row = self._write_section_header(ws, row, "OTURUM BİLGİLERİ")
        
        # Calculate duration text
        if session.duration_seconds:
            minutes = session.duration_seconds // 60
            seconds = session.duration_seconds % 60
            duration_text = f"{minutes} dakika {seconds} saniye"
        else:
            duration_text = "---"
        
        session_data = [
            ("Başlangıç:", session.start_time.strftime(config.DATETIME_FORMAT) if session.start_time else "---"),
            ("Bitiş:", session.end_time.strftime(config.DATETIME_FORMAT) if session.end_time else "Devam Ediyor"),
            ("Süre:", duration_text),
            ("Tamamlanma:", f"%{session.get_completion_percentage():.0f}"),
            ("Başarılı:", str(session.get_passed_count())),
            ("Başarısız:", str(session.get_failed_count())),
        ]
        
        for label, value in session_data:
            row = self._write_info_row(ws, row, label, value)
        
        return row
    
    def _write_section_header(self, ws, row: int, title: str) -> int:
        """Write a section header row"""
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws[f'A{row}']
        cell.value = title
        cell.font = self.SECTION_FONT
        cell.fill = self.SECTION_FILL
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = self.BORDER_THIN
        return row + 1
    
    def _write_info_row(self, ws, row: int, label: str, value: str) -> int:
        """Write a label-value info row"""
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].border = self.BORDER_THIN
        
        ws.merge_cells(f'B{row}:G{row}')
        ws[f'B{row}'] = value
        ws[f'B{row}'].border = self.BORDER_THIN
        
        return row + 1
    
    def _write_calibration_row(self, ws, row: int, label: str, date_value: str, status: str) -> int:
        """Write a calibration row with status-based coloring"""
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].border = self.BORDER_THIN
        
        ws.merge_cells(f'B{row}:F{row}')
        ws[f'B{row}'] = date_value
        ws[f'B{row}'].border = self.BORDER_THIN
        
        # Status indicator with color
        status_cell = ws[f'G{row}']
        if status == 'expired':
            status_cell.value = "SÜRESİ DOLMUŞ"
            status_cell.fill = self.EXPIRED_FILL
            status_cell.font = Font(color="B71C1C", bold=True)
        elif status == 'warning':
            status_cell.value = "YAKINDA DOLACAK"
            status_cell.fill = self.WARNING_FILL
            status_cell.font = Font(color="E65100", bold=True)
        else:
            status_cell.value = "GEÇERLİ"
            status_cell.fill = self.PASS_FILL
            status_cell.font = Font(color="1B5E20", bold=True)
        
        status_cell.border = self.BORDER_THIN
        status_cell.alignment = Alignment(horizontal='center')
        
        return row + 1
    
    def _write_results_table(self, ws, session: TestSession, start_row: int) -> int:
        """
        Write test results table with step details.
        
        Returns:
            Next available row number
        """
        row = start_row
        
        # Table header
        row = self._write_section_header(ws, row, "TEST SONUÇLARI")
        
        # Table column headers
        headers = [
            "Adım No",
            "Adım Adı",
            "Durum",
            "Başlangıç",
            "Süre (sn)",
            "Sonuç",
            "Yorum"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER_THIN
        
        row += 1
        
        # Table data
        for step in session.steps:
            col_idx = 1
            
            # Step ID
            cell = ws.cell(row=row, column=col_idx)
            cell.value = step.step_id
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER_THIN
            col_idx += 1
            
            # Step Name
            cell = ws.cell(row=row, column=col_idx)
            cell.value = step.name
            cell.border = self.BORDER_THIN
            col_idx += 1
            
            # Status (with color coding)
            cell = ws.cell(row=row, column=col_idx)
            status_text, status_fill = self._get_status_display(step.status)
            cell.value = status_text
            cell.fill = status_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER_THIN
            col_idx += 1
            
            # Start Time
            cell = ws.cell(row=row, column=col_idx)
            if step.start_time:
                cell.value = datetime.fromtimestamp(step.start_time).strftime(config.DATETIME_FORMAT)
            else:
                cell.value = "---"
            cell.border = self.BORDER_THIN
            col_idx += 1
            
            # Duration
            cell = ws.cell(row=row, column=col_idx)
            cell.value = step.actual_duration if step.actual_duration is not None else "---"
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER_THIN
            col_idx += 1
            
            # Result Value
            cell = ws.cell(row=row, column=col_idx)
            cell.value = self._format_result_value(step.result_value)
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER_THIN
            col_idx += 1
            
            # Comment
            cell = ws.cell(row=row, column=col_idx)
            cell.value = step.comment if step.comment else ""
            cell.border = self.BORDER_THIN
            col_idx += 1
            
            row += 1
        
        return row
    
    def _get_status_display(self, status: TestStatus) -> tuple:
        """
        Get display text and fill color for status.
        
        Returns:
            Tuple of (status_text, fill_color)
        """
        if status == TestStatus.PASSED:
            return ("BAŞARILI", self.PASS_FILL)
        elif status == TestStatus.FAILED:
            return ("BAŞARISIZ", self.FAIL_FILL)
        elif status == TestStatus.IN_PROGRESS:
            return ("DEVAM EDİYOR", PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"))
        elif status == TestStatus.SKIPPED:
            return ("ATLANDI", self.NA_FILL)
        else:
            return ("BAŞLANMAMIŞ", self.NA_FILL)
    
    def _format_result_value(self, value) -> str:
        """Format result value for display"""
        if value is None:
            return "---"
        elif value in ["PASS", "GEÇTİ"]:
            return "GEÇTİ"
        elif value in ["FAIL", "KALDI"]:
            return "KALDI"
        else:
            return str(value)
    
    def _auto_size_columns(self, ws):
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            # Set width with some padding
            adjusted_width = min(max_length + 2, 50)  # Max 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def generate_filename(self, session: TestSession, output_dir: str) -> str:
        """
        Generate Excel filename based on session data.
        
        Format: TestRaporu_<DATE>_<STOCK>_<SERIAL>.xlsx
        
        Args:
            session: TestSession object
            output_dir: Directory where file will be saved
            
        Returns:
            Full path to output file
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        stock = self._sanitize_filename(session.stock_number)
        serial = self._sanitize_filename(session.serial_number)
        
        filename = f"TestRaporu_{timestamp}_{stock}_{serial}.xlsx"
        
        return str(Path(output_dir) / filename)
    
    def _sanitize_filename(self, name: str) -> str:
        """Remove invalid characters from filename"""
        if not name:
            return "unknown"
        # Keep only alphanumeric, hyphens, and underscores
        return ''.join(c if c.isalnum() or c in '-_' else '_' for c in name)